# -*- coding: utf-8 -*-
try:

    import os
    import sys
    import traceback
    import datetime
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from getpass import getpass
    import subprocess
    import json
    from atlassian import Confluence
    from atlassian.errors import ApiValueError
    from copy import copy
    from openpyxl.utils.cell import range_boundaries
    
    import logging
    import warnings
    
    import sys
    import locale
    import time
    from requests import HTTPError
    
    from progress.bar import IncrementalBar
    
    warnings.simplefilter(action='ignore', category=UserWarning)
    start_date = datetime.datetime.today()
    print(f'Начало выполнения: {start_date}')
    
    log = logging.getLogger(__name__)
        
    class ConfluenceOver(Confluence):
        content_types = {
        ".gif": "image/gif",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".pdf": "application/pdf",
        ".doc": "application/msword",
        ".xls": "application/vnd.ms-excel",
        ".svg": "image/svg+xml",
    }
        def get_page_by_id(self, page_id, expand=None, status=None, version=None):
            """
            Returns a piece of Content.
            Example request URI(s):
            http://example.com/confluence/rest/api/content/1234?expand=space,body.view,version,container
            http://example.com/confluence/rest/api/content/1234?status=any
            :param page_id: Content ID
            :param status: (str) list of Content statuses to filter results on. Default value: [current]
            :param version: (int)
            :param expand: OPTIONAL: Default value: history,space,version
                           We can also specify some extensions such as extensions.inlineProperties
                           (for getting inline comment-specific properties) or extensions.resolution
                           for the resolution status of each comment in the results
            :return:
            """
            params = {}
            if expand:
                params["expand"] = expand
            if status:
                params["status"] = status
            if version:
                params["version"] = version
            url = "rest/api/content/{page_id}".format(page_id=page_id)

            try:
                response = self.get(url, params=params)
                time.sleep(1)
            except HTTPError as e:
                if e.response.status_code == 404:
                    # Raise ApiError as the documented reason is ambiguous
                    raise ApiError(
                        "There is no content with the given id, "
                        "or the calling user does not have permission to view the content",
                        reason=e,
                    )

                raise

            return response
        def update_page(
        self,
        page_id,
        title,
        body=None,
        parent_id=None,
        type="page",
        representation="storage",
        minor_edit=False,
        version_comment=None,
        always_update=False,
        ):
            """
            Update page if already exist
            :param page_id:
            :param title:
            :param body:
            :param parent_id:
            :param type:
            :param representation: OPTIONAL: either Confluence 'storage' or 'wiki' markup format
            :param minor_edit: Indicates whether to notify watchers about changes.
                If False then notifications will be sent.
            :param version_comment: Version comment
            :param always_update: Whether always to update (suppress content check)
            :return:
            """
            log.info('Updating {type} "{title}"'.format(title=title, type=type))
            if not always_update and body is not None and self.is_page_content_is_already_updated(page_id, body, title):
                return self.get_page_by_id(page_id)

            try:
                if self.advanced_mode:
                    version = self.history(page_id).json()["lastUpdated"]["number"] + 1
                    time.sleep(1)
                else:
                    version = self.history(page_id)["lastUpdated"]["number"] + 1
                    time.sleep(1)
            except (IndexError, TypeError) as e:
                log.error("Can't find '{title}' {type}!".format(title=title, type=type))
                log.debug(e)
                return None

            data = {
                "id": page_id,
                "type": type,
                "title": title,
                "version": {"number": version, "minorEdit": minor_edit},
            }
            if body is not None:
                data["body"] = self._create_body(body, representation)
                time.sleep(1)

            if parent_id:
                data["ancestors"] = [{"type": "page", "id": parent_id}]
            if version_comment:
                data["version"]["message"] = version_comment

            try:
                response = self.put("rest/api/content/{0}".format(page_id), data=data)
                time.sleep(1)
            except HTTPError as e:
                if e.response.status_code == 400:
                    raise ApiValueError(
                        "No space or no content type, or setup a wrong version "
                        "type set to content, or status param is not draft and "
                        "status content is current",
                        reason=e,
                    )
                if e.response.status_code == 404:
                    raise ApiNotFoundError("Can not find draft with current content", reason=e)

                raise
            return response
        
        
        def attach_content(
        self,
        content,
        name,
        content_type="application/binary",
        page_id=None,
        title=None,
        space=None,
        comment=None,
        ):
            """
            Attach (upload) a file to a page, if it exists it will update the
            automatically version the new file and keep the old one.
            :param title: The page name
            :type  title: ``str``
            :param space: The space name
            :type  space: ``str``
            :param page_id: The page id to which we would like to upload the file
            :type  page_id: ``str``
            :param name: The name of the attachment
            :type  name: ``str``
            :param content: Contains the content which should be uploaded
            :type  content: ``binary``
            :param content_type: Specify the HTTP content type. The default is
            :type  content_type: ``str``
            :param comment: A comment describing this upload/file
            :type  comment: ``str``
            """
            page_id = self.get_page_id(space=space, title=title) if page_id is None else page_id
            time.sleep(1)
            type = "attachment"
            if page_id is not None:
                comment = comment if comment else "Uploaded {filename}.".format(filename=name)
                time.sleep(1)
                data = {
                    "type": type,
                    "fileName": name,
                    "contentType": content_type,
                    "comment": comment,
                    "minorEdit": "true",
                }
                time.sleep(1)
                headers = {"X-Atlassian-Token": "no-check", "Accept": "application/json"}
                time.sleep(1)
                path = "rest/api/content/{page_id}/child/attachment".format(page_id=page_id)
                time.sleep(1)
                # Check if there is already a file with the same name
                attachments = self.get(path=path, headers=headers, params={"filename": name})

                if attachments.get("size"):
                    time.sleep(1)
                    path = path + "/" + attachments["results"][0]["id"] + "/data"

                try:
                    time.sleep(1)
                    response = self.post(
                        path=path,
                        data=data,
                        headers=headers,
                        files={"file": (name, content, content_type)},
                    )
                    time.sleep(1)
                except HTTPError as e:
                    if e.response.status_code == 403:
                        # Raise ApiError as the documented reason is ambiguous
                        raise ApiError(
                            "Attachments are disabled or the calling user does "
                            "not have permission to add attachments to this content",
                            reason=e,
                        )
                    if e.response.status_code == 404:
                        # Raise ApiError as the documented reason is ambiguous
                        raise ApiError(
                            "The requested content is not found, the user does not have "
                            "permission to view it, or the attachments exceeds the maximum "
                            "configured attachment size",
                            reason=e,
                        )

                    raise
                return response
            else:
                log.warning("No 'page_id' found, not uploading attachments")
                return None
               
    login = os.getlogin()
    pswd = getpass(prompt="Password:")
    
    confluence = ConfluenceOver(
        url=<УДалено специально>
        username=login,
        password= pswd,
        verify_ssl=False
    )
    
    
    def update_page_info(page):
        json_page = confluence.get_page_by_id(page_id=page, expand='body.view')
        time.sleep(1)
        body_value = str(json_page['body']['view']['value'])
        body_index = body_value.find('202')
        body_value = body_value[:body_index-6] + str(datetime.date.today().strftime('%d.%m.%Y')) + body_value[body_index+4:]
        input(body_value)
        title_ = json_page['title']
        confluence.update_page(page_id = page, title = title_, body = body_value)
    
    
    def attach(file, page):
        confluence.attach_file(filename=file, page_id=page)
    def del_attach(file, page):
        confluence.delete_attachment(filename=file, page_id=page)

    
    
    
    
    cwd = os.getcwd()
    files = os.listdir()
    list_corp = [<тут список с названиями блоков, удалено специально>]
    list_pageId = [<тут список с id страниц в CONFluence, удалено специально>]
    list_conf_names = [<Названия файлов по блокам для COnfluence, удалено специально>]

    folder = str(datetime.date.today())
    if (os.path.exists('../' + folder) == False):
        os.mkdir('../' + folder)
    
    
    start_page = input('Номер страницы:')
    iterator = input('Задать шаблон(0), отмена(1):')
    if (iterator != '1'):
        iterator = 0
        print(f'{iterator}: будет задан новый шаблон')
    else:
        iterator = int(iterator)
        print(f'{iterator}: будет использован имеющийся шаблон')
    
    def del_files(sub_directory, area):
        name = '../' +folder+'/' + sub_directory
        files = os.listdir(name)
        for file in files:
            time.sleep(1)
            attachments = confluence.get_attachments_from_content(page_id = list_pageId[area], start=0, limit=50, expand=None, filename=file,
                                                              media_type=None)
            if (len(attachments['results']) > 0): 
                try:
                    del_attach(file, list_pageId[area])
                    time.sleep(1)
                    for res in attachments['results']:
                        if res['title'] == file:
                            confluence.remove_page_from_trash(res['id'])
                except Exception as error:
                    continue

    #Пока не обновляется page_info, нужно настроить
    def up_files_and_info(sub_directory, area):
        name = '../' +folder+'/' + sub_directory + '/'
        files = os.listdir(name)
        for file in files:
            file = name + file
            print(file)
            attach(file, list_pageId[area])
            time.sleep(1)
        #update_page_info(list_pageId[area])
        time.sleep(1)
    
    
    upload_flag = input("Для загрузки файлов на Confluence наберите 'yes' (при условии, что файлы уже обновлены):\n")
    if (upload_flag == 'yes'):
        for area in range(int(start_page), len(list_corp), 1):        
            del_files(list_conf_names[area], area)
            up_files_and_info(list_conf_names[area], area)
            if(area == len(list_corp) - 1):
                input('finish')
                sys.exit()
        
    
    
    temp_name = '../' +folder+'/' + 'temp'
    for area in range(int(start_page), len(list_corp), 1):
        print('###################################################')
        print(str(area+1) + '. ' + list_corp[area] + '\nloading...')

        district = list_corp[area]
        block = district.upper()
        block_compare_name = list_conf_names[area] + '.xlsx'
        name = '../' +folder+'/' + list_conf_names[area]
        #Создаем директорию под каждый блок
        if (os.path.exists(name) == False):
            os.mkdir(name)
        name = '../' +folder+'/' + list_conf_names[area] + '/' + list_conf_names[area]
        
        #print(name)
        template = '../template'

        #Функция проверки названия листа
        def sheet_name(list):
            for area in list:
                if str(area).lower().__contains__("реестр"): #добавить raise error если нет такого листа
                    return area
        #Копирование заголовка из каждого файла в выходной на отдельную страницу
        
        def copy_header_to_file(input_dict, temp):
            excel_file = Workbook()                        
            del excel_file['Sheet']
            #проход по всем файлам словаря;
            for key in input_dict.keys():
                #Загрузка книги, которую будем обрабатывать(из нее будет взята шапка)
                wb = load_workbook(filename = key)#'2019_full.xlsx')
                #Загрузка листа в названии которого есть слово "реестр"
                for i in wb.sheetnames: 
                    if ((str(i).lower() == ('реестр'))):
                        registry_sheet = wb[i] #Нужно добавить выброс из функции
                        wb.close() #файл из которого забирали данные и шапку
                '''
                Выход на следующий файл, если не найдено строк по нужному блоку
                Сделано для того, чтобы не создавать лишние пустые страницы
                '''
                # df = pd.DataFrame(registry_sheet.values)
                # df.columns = df.iloc[1]
                # df = df.drop(index=[0,1])
                # try:
                   # df = df.loc[(df['Блок заказчика'].str.contains(block, case=False))]
                # except Exception as error: #Обычно это ошибка пустых значений в столбце "Блок заказчика"
                   # df['Блок заказчика'] = df['Блок заказчика'].fillna('')  
                   # df = df.loc[(df['Блок заказчика'].str.contains(block, case=False))]
                # if(len(df) == 0):
                   # continue
                '''
                Окончание кода:
                Выход на следующий файл, если не найдено строк по нужному блоку
                Сделано для того, чтобы не создавать лишние пустые страницы
                '''
                header_cells_generator = registry_sheet.iter_rows(max_row=2)

                #Каждый файл добавляем на новый лист
                excel_sheet = excel_file.create_sheet(title=input_dict[key])

                #Копирование объединеннных ячеек в выходной excel file
                for _range in registry_sheet.merged_cells.ranges:
                    boundaries = range_boundaries(str(_range))
                    excel_sheet.merge_cells(start_column=boundaries[0], start_row=boundaries[1],
                            end_column=boundaries[2], end_row=boundaries[3])
                    
                #копирование значений шапки и стиля ячеек в выходной файл
                orig_row = 0
                for header_cells_tuple in header_cells_generator:
                    orig_row +=1
                    excel_sheet.row_dimensions[orig_row].height = 20
                    for i in range(len(header_cells_tuple)):
                        excel_sheet.cell(row=orig_row, column=i+1, value = header_cells_tuple[i].value)
                        excel_sheet.cell(row=orig_row, column=i+1).font = copy(header_cells_tuple[i].font)
                        excel_sheet.cell(row=orig_row, column=i+1).border = copy(header_cells_tuple[i].border)
                        excel_sheet.cell(row=orig_row, column=i+1).fill = copy(header_cells_tuple[i].fill)
                        excel_sheet.cell(row=orig_row, column=i+1).number_format = copy(header_cells_tuple[i].number_format)
                        excel_sheet.cell(row=orig_row, column=i+1).protection = copy(header_cells_tuple[i].protection)
                        excel_sheet.cell(row=orig_row, column=i+1).alignment = copy(header_cells_tuple[i].alignment)
            excel_file.save(filename=temp)
            excel_file.close()
        
        #Функция добавления данных в файл (шапка создается в функции copy_header_to_file)
        def data_to_file(input_dict, temp, load_file):
            book = load_workbook(temp)
            book.save(filename=load_file)
            book.close()
            with pd.ExcelWriter(load_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:   
                writer.book.create_sheet(title='No orders')
                for key in input_dict.keys():
                    file = str(key)
                    xl = pd.ExcelFile(file)
                    data = xl.parse(sheet_name(xl.sheet_names), header=1)  
                    try:
                        data = data.loc[(data['Блок заказчика'].str.contains(block, case=False))]
                        if(len(data) == 0):
                            writer.book.remove(writer.book[input_dict[file]])  
                            continue
                    except Exception as error: #Обычно это ошибка пустых значений в столбце "Блок заказчика"
                        data['Блок заказчика'] = data['Блок заказчика'].fillna('')
                        data = data.loc[(data['Блок заказчика'].str.contains(block, case=False))]
                        if(len(data) == 0):
                            writer.book.remove(writer.book[input_dict[file]])
                            continue
                    xl.close()             
                    data.to_excel(writer, input_dict[file], index=False, startrow=2, header=False)
                if(len(writer.book.sheetnames) > 1):
                    writer.book['No orders'].sheet_state = 'hidden'
                    # with  as writer:
                            # data.to_excel(writer, input_dict[file], index=False, startrow=2, header=False)   

                            
        <НАзвания удалены специально>
        #Словарь названий файлов и названий листов для будущего файла 2019 года
        dict_2019 = {<НАзвания удалены специально>}
                
        <НАзвания удалены специально>
        #Словарь названий файлов и названий листов для будущего файла 2020 года
        dict_2020 = {<НАзвания удалены специально>}

        
        <НАзвания удалены специально>
        #Словарь названий файлов и названий листов для будущего файла 2021 года
        dict_2021 = {<НАзвания удалены специально>
            }

        <НАзвания удалены специально>
        #Словарь названий файлов и названий листов для будущего файла 2022 года
        dict_2022 = {<НАзвания удалены специально>}
        
        
        general_dict = {1:dict_2019, 2:dict_2020, 3:dict_2021, 4:dict_2022}
       
        bar = IncrementalBar('Обновление файлов', max = len(general_dict))
        bar.next(0)
        file_year = 2019    
        for dictionary_key in general_dict:                  
            load_file = name + str(file_year) + '.xlsx'
            temp = temp_name + str(file_year) + '.xlsx'
            #Копирование заголовка
            if (iterator == 0):
                copy_header_to_file(general_dict[dictionary_key], temp)
            #Кусок добавления самих данных блока
            data_to_file(general_dict[dictionary_key], temp, load_file)
            bar.next()
            file_year += 1
        
            
        bar.finish()
        end_date = datetime.datetime.today()
        
        diff = end_date - start_date
        print(f'{start_date}\n{end_date}\n{diff}')
        iterator += 1
    input('Все файлы обновлены\npress Enter...')       

except Exception as error:
    print(error)
    print(sys.exc_info())
    print(traceback.format_exc())
    input("Update Failed, press Enter...")